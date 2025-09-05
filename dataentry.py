from PyQt5.QtGui import *
from PyQt5.QtCore import *
from PyQt5.QtWidgets import *
from PyQt5.uic import loadUi 
import sys
from openpyxl import Workbook , load_workbook
import sqlite3
from openpyxl.styles import NamedStyle, Font, Border, Side
from openpyxl.utils import get_column_letter
import os
class data_window (QWidget):
    def __init__(self):
        super(data_window,self).__init__()
        loadUi("data_win.ui",self)
        
        self.tabel1.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table2.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table3.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table4.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.file=self.file_name1.text() 
        self.sheet_name=self.sheet1.text()
        self.creatf_btn.clicked.connect(self.creat_file)
        self.savedata1.clicked.connect(self.savedata11)
        self.browsebtn.clicked.connect(self.browse_file)
        self.save2btn.clicked.connect(self.savedata)
        self.browsebtn_2.clicked.connect(self.browse2)
        self.viewbtn.clicked.connect(self.view_data)
        self.browsebtn3.clicked.connect(self.browse3)
        self.search_btn.clicked.connect(self.search)
        self.cleartabel.clicked.connect(self.cleardata)
        layout = QVBoxLayout()
        layout.addWidget(self.file_name1)
        layout.addWidget(self.sheet1)
        layout.addWidget(self.creatf_btn)
        # ضيف كل العناصر بالشكل اللي ترتبه بيه

        layout.addWidget(self.tabel1)  # خلي الجدول في الآخر عشان يتمدد كويس

        self.setLayout(layout)
        self.f=0
        self.f2=0
    def creat_file(self):
        self.file_name1.text()
        f_name = self.file_name1.text() + ".xlsx"
        sheet_name = self.sheet1.text()
        if os.path.exists(f_name):
            m = QMessageBox.question(self, "Error", "This File Is Already exist ,Do you want to delet it and make new one ?",QMessageBox.Yes | QMessageBox.No)
            if m == QMessageBox.Yes:
                wb=Workbook()
                ws=wb.create_sheet(title=sheet_name)
                wb.save(f_name)
                main_row=['Name','Birth Date','Address','Job','Salary','Notes']
                ws.append(main_row)
                wb.save(f_name)
                QMessageBox.about(self, "Done", "Sheet Is created ;Start adding Data")
            else :
                self.file_name1.setText("")
                self.sheet1.setText("")
                QMessageBox.about(self, "OK", "Choose another file name or go to second tan and browse file ")
        else:
            wb = Workbook()
            ws = wb.create_sheet(title=sheet_name)
            wb.save(f_name)
            main_row = ['Name', 'Birth Date', 'Address', 'Job', 'Salary', 'Notes']
            ws.append(main_row)
            wb.save(f_name)
            QMessageBox.about(self, "Done", "Sheet Is created ;Start adding Data")
    def savedata11(self):
      #self.f=0 
      f_name=self.file_name1.text()+ ".xlsx"
      sheet=self.sheet1.text()
      data = []
      if self.f>= self.tabel1.rowCount():
           self.tabel1.setRowCount(self.f +1)
    # إنشاء كائنات QTableWidgetItem لكل قيمة
      name_item = QTableWidgetItem(self.nam2_ent.text())
      birth_date_item = QTableWidgetItem(self.birth2.text())
      address_item = QTableWidgetItem(self.add2_ent.text())
      jop_item = QTableWidgetItem(self.jop2_ent.text())
      sallary_item = QTableWidgetItem(str(self.sal2_ent.text()))  # تحويل الرقم إلى نص
      notes_item = QTableWidgetItem(self.not2_ent.text())
    # إضافة العناصر إلى الجدول باستخدام setItem
      self.tabel1.setItem(self.f, 0, name_item)
      self.tabel1.setItem(self.f, 1, birth_date_item)
      self.tabel1.setItem(self.f, 2, address_item)
      self.tabel1.setItem(self.f, 3, jop_item)
      self.tabel1.setItem(self.f, 4, sallary_item)
      self.tabel1.setItem(self.f, 5, notes_item)
      self.f +=1
    # إضافة البيانات إلى القائمة (data) للحفظ في الملف
      data.append(self.nam2_ent.text())
      data.append(self.birth2.text())
      data.append(self.add2_ent.text())
      data.append(self.jop2_ent.text())
      data.append(self.sal2_ent.text())
      data.append(self.not2_ent.text())

    # باقي الكود كما هو
      wb = load_workbook(f_name)
      ws = wb[sheet]
      ws.append(data)
      #self.sheet_style(f_name,sheet)
      wb.save(f_name)
      self.nam2_ent.clear()
      self.birth2.clear()
      self.add2_ent.clear()
      self.add2_ent.clear()
      self.jop2_ent.clear()
      self.sal2_ent.clear()
      self.not2_ent.clear()
    def browse_file(self) :
        self.file_name.clear()
        self.sheet_combo.clear()
        dialog = QFileDialog(self)
        dialog.setFileMode(QFileDialog.ExistingFile)  # اختيار ملف واحد
        dialog.setNameFilter('EXCEL files (*.xlsx)')  # فلتر لأنواع الملفات

        # فتح الديالوج وفحص النتيجة
        if dialog.exec_() == QFileDialog.Accepted:
            selected_file = dialog.selectedFiles()[0]
            self.file_name.setText(selected_file)
            wb=load_workbook(selected_file)
            sheets=wb.sheetnames
            for sheet in sheets:
                self.sheet_combo.addItem(sheet)
        else:
            # لو ضغط Cancel
            QMessageBox.about(self, "File Canseled", "You canceled choose file")
    def savedata(self):  
      data1=[]

      if self.f2>= self.table2.rowCount():
           self.table2.setRowCount(self.f2 +1)
      f_name=self.file_name.text()
      sh_name=self.sheet_combo.currentText()
      name_item = QTableWidgetItem(self.name_entry.text())
      birth_date_item = QTableWidgetItem(self.birth_entry.text())
      address_item = QTableWidgetItem(self.add_ent.text())
      jop_item = QTableWidgetItem(self.jop_ent.text())
      sallary_item = QTableWidgetItem(str(self.sall_ent.text()))  # تحويل الرقم إلى نص
      notes_item = QTableWidgetItem(self.note_ent.text())
    # إضافة العناصر إلى الجدول باستخدام setItem
      self.table2.setItem(self.f2, 0, name_item)
      self.table2.setItem(self.f2, 1, birth_date_item)
      self.table2.setItem(self.f2, 2, address_item)
      self.table2.setItem(self.f2, 3, jop_item)
      self.table2.setItem(self.f2, 4, sallary_item)
      self.table2.setItem(self.f2, 5, notes_item)
      self.f2 +=1
    # إضافة البيانات إلى القائمة (data) للحفظ في الملف
      data1.append(self.name_entry.text())
      data1.append(self.birth_entry.text())
      data1.append(self.add_ent.text())
      data1.append(self.jop_ent.text())
      data1.append(self.sall_ent.text())
      data1.append(self.note_ent.text())
      wb = load_workbook(f_name)
      ws = wb[sh_name]
      ws.append(data1)
     # self.sheet_style(f_name,sh_name)
      wb.save(f_name)
      self.name_entry.clear()
      self.birth_entry.clear()
      self.add_ent.clear()
      self.jop_ent.clear()
      self.sall_ent.clear()
      self.note_ent.clear()
    def browse2(self):
        dialog = QFileDialog(self)
        dialog.setFileMode(QFileDialog.ExistingFile)  # اختيار ملف واحد
        dialog.setNameFilter('EXCEL files (*.xlsx)')  # فلتر لأنواع الملفات

        # فتح الديالوج وفحص النتيجة
        if dialog.exec_() == QFileDialog.Accepted:
            # لو اختار ملف (Accept)
            selected_file = dialog.selectedFiles()[0]
            self.search_file.setText(selected_file)
            wb = load_workbook(selected_file)
            sheets = wb.sheetnames
            for sheet in sheets:
                self.sheet_combo_2.addItem(sheet)
        else:
            # لو ضغط Cancel
            QMessageBox.about(self, "File Canseled", "You canceled choose file")
            # هنا مابنعملش حاجة، البرنامج هيفضل مفتوح


    def view_data (self):
        v_file= self.search_file.text()
        wb=load_workbook(v_file)
        name=self.sheet_combo_2.currentText()
        ws=wb[name]
        r_max=ws.max_row +1
        c_max=ws.max_column +1
        self.table4.setRowCount(r_max)  
        #self.tableWidget.setColumnCount(c_max)  
        self.row=0
        for i in range (1,r_max):
           data=[]
           for j in range (1,c_max):
               val=ws.cell(row=i,column=j).value
               data.append(val)
           self.row=0
           for e in range(0,(c_max-1)):
              self.table4.setItem(i, e, QTableWidgetItem(str(data[e])))
           self.row+=1
    def browse3(self):
        self.file_name2.clear()
        self.sheet_combo2.clear()
        dialog = QFileDialog(self)
        dialog.setFileMode(QFileDialog.ExistingFile)  # اختيار ملف واحد
        dialog.setNameFilter('EXCEL files (*.xlsx)')  # فلتر لأنواع الملفات

        # فتح الديالوج وفحص النتيجة
        if dialog.exec_() == QFileDialog.Accepted:
            # لو اختار ملف (Accept)
            selected_file = dialog.selectedFiles()[0]
            self.file_name2.setText(selected_file)
            wb=load_workbook(selected_file)
            sheets=wb.sheetnames
            for sheet in sheets:
              self.sheet_combo2.addItem(sheet)
        else:
               # لو ضغط Cancel
               QMessageBox.about(self, "File Canseled", "You canceled choose file")
    def search(self):
        s_file= self.file_name2.text()
        wb=load_workbook(s_file)
        name=self.sheet_combo2.currentText()
        ws=wb[name]
        r_max=ws.max_row +1
        c_max=ws.max_column +1
        #self.table3.setColumnCount(c_max)  
        self.table3.setRowCount(r_max)  
        self.row=0
        searchword=self.name_entry_2.text()
        count=0
        for i in range (1,r_max):
           data=[]
           for j in range (1,c_max):
               val=ws.cell(row=i,column=j).value
               if str(val) ==searchword:
                   e=i
                   count +=1
                   data = []
                   for a in range (1,c_max):
                      v=ws.cell(row=e,column=a).value
                      data.append(str(v))

                   for d in range(0,(c_max-1)):
                     self.table3.setItem(self.row, d, QTableWidgetItem(str(data[d])))
                   self.row+=1
                   data = []
        if count == 0 :
            QMessageBox.about(self, "Not Found", "Word not found ,check spelling and try again")
    def cleardata(self):
        self.table3.clearContents()
        self.name_entry_2.setText("")
        self.sheet_combo2.clear()
        self.file_name2.setText("")
    def sheet_style(self,file_name,sheet):
        highlight = NamedStyle(name="highlight")
        highlight.font = Font(bold=True, size=20)
        bd = Side(style='thin', color="000000")
        highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        wb = load_workbook(file_name)
        ws = wb[sheet]
        r_max = ws.max_row + 1
        r_min = ws.min_row
        c_max = ws.max_column + 1
        c_min = ws.min_column
        for i in range(r_min, r_max):
            for j in range(c_min, c_max):
                ws.cell(row=i, column=j).style = highlight
        for i in range(r_min, r_max):
            for j in range(c_min, c_max):
                val = ws.cell(row=i, column=j).value
                colum = j
                colet = get_column_letter(colum)
                ll = ws.column_dimensions[colet].width
                if ll < len(str(val)):
                    ll == len(str(val)) + 2
        wb.save(file_name)


class signup_window (QWidget):
    def __init__(self):
        super(signup_window,self).__init__()
        loadUi("secwin.ui",self)
        self.dic={}
        self.gen=""
        self.save_newuser.clicked.connect(self.save_user)
    def maleselected(self, selected):
        if selected:
            self.gen= "Male"
        return self.gen
    def femaleselected(self, selected):
        if selected:
            self.gen= "Female"
        return self.gen
    def save_user(self):
        new_user=[]
        name=self.new_user_name.text()
        pas=self.new_pass.text()
        self.male_rad.toggled.connect(self.maleselected)
        self.fe_rad.toggled.connect(self.femaleselected)
        if self.male_rad.isChecked():
             self.gen="Male"
        if self.fe_rad.isChecked():
             self.gen="Female"
        if name =="" or pas=="" or self.gen=="" :
            QMessageBox.about(self, "Error", "PLEASE ENTER YOUR DATA FOR SIGN IN") 
        else:
         """wb=load_workbook("work.xlsx")
         ws=wb["Sheet"]
         new_user.append(name)
         new_user.append(pas)
         new_user.append(self.gen)
         ws.append(new_user)
         wb.save("work.xlsx")
         QMessageBox.about(self, "Congratulation", "You are signed up sucssefully") 
         self.close()
         window.show()"""
         con=sqlite3.connect("users.db")
         cursor=con.cursor()

         dic={}
         for row in db :
             dic[row[1]]=row[2]
         for key in dic:
           if name == key :
              QMessageBox.about(self, "Error", "This name is already existing ,Please choose other name or cancel and Sign IN") 
           else:
             cursor.execute("INSERT INTO users VALUES(?,?,?)",(name,pas,self.gen))
         con.commit()
         con.close()
         QMessageBox.about(self, "Congratulation", "You are signed up sucssefully") 
         self.close()
         window.show()
         print(dic)
class login (QWidget):
    def __init__(self):
        super(login,self).__init__()
        loadUi("login.ui",self)
        self.sign_btn.clicked.connect(self.sign_in)
        self.signup_btn.clicked.connect(self.sign_up)
        self.user=self.user_entry_2.text()
        self.pass_entry.setEchoMode(QLineEdit.Password)
    def sign_in(self):
       user=self.user_entry_2.text()
       pass1=self.pass_entry.text()
       if user =="" or pass1=="" :
            QMessageBox.about(self, "Error", "PLEASE ENTER YOUR DATA FOR SIGN IN") 
       else:    
         """  wb=load_workbook("work.xlsx")
         ws=wb["Sheet"]
         dic={}
         r_max=ws.max_row +1
         for i in range (2,r_max):
          user_name=ws.cell(row=i,column=1).value
          user_p=ws.cell(row=i,column=2).value
          user_pass=str(user_p)
          dic[user_name]=user_pass
         if user in dic.keys() and dic[user]==pass1:
              self.data_window1=data_window()
              self.data_window1.show()
              self.close()
         else:
             QMessageBox.about(self, "Error", "Wrong user name or password")"""
         con=sqlite3.connect("users.db")
         cursor=con.cursor()
         db=cursor.execute("SELECT rowid,* FROM users ")
         dic={}
         for row in db :
             dic[row[1]]=row[2]
         if user in dic and dic[user]==pass1: 
              self.data_window1=data_window()
              self.data_window1.show()
              self.close()
         else:
             QMessageBox.about(self, "Error", "Wrong user name or password")
         con.commit()
         con.close()
    def sign_up(self):
          self.sec_window1=signup_window()
          self.sec_window1.show()
          self.close()
        
if __name__=="__main__":
    app=QApplication(sys.argv)
    window=login()
    window.show()
    sys.exit(app.exec())  